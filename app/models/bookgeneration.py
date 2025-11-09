import os
import asyncio
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from typing import Annotated, List, Optional, Dict, Any
from typing_extensions import TypedDict
from langchain_core.messages import HumanMessage
from langchain_core.prompts import ChatPromptTemplate
from langchain_google_genai import ChatGoogleGenerativeAI
from langgraph.graph import StateGraph, START, END
from langgraph.graph.message import add_messages
from langgraph.checkpoint.memory import InMemorySaver
import openpyxl
from dotenv import load_dotenv
from app.utils.db_getter import get_db

load_dotenv(override=True)


class State(TypedDict):
    messages: Annotated[list, add_messages]
    title: str
    sanitized_title: str
    notes_before_outline: str
    outline: str
    notes_on_outline_after: Optional[str]
    chapters: List[str]
    chapter_summaries: List[Dict[str, Any]]  
    current_chapter_number: int
    user_decision: str
    summary: str
    chapter_feedback: Optional[str]
    final_review_notes_status: Optional[str]
    final_review_notes: Optional[str]


class BookGeneration:
    def __init__(self, excel_fp ):
        self.llm = ChatGoogleGenerativeAI(model="gemini-2.5-flash")
        self.excel_filename = excel_fp
        self.db = get_db()  # Initialize database client
        
        self.sender_email = os.getenv("SENDER_EMAIL")
        self.email_password = os.getenv("EMAIL_PASSWORD")
        self.receiver_email = os.getenv("RECEIVER_EMAIL")

        graph_builder = StateGraph(State)

        # Define the nodes
        graph_builder.add_node("read_local_excel", self.read_local_excel)
        graph_builder.add_node("generate_outline", self.generate_outline_node)
        graph_builder.add_node("write_outline_to_excel", self.write_outline_to_excel)
        graph_builder.add_node("human_feedback_interrupt", self.human_feedback_interrupt)
        graph_builder.add_node("read_feedback_from_excel", self.read_feedback_from_excel)
        graph_builder.add_node("generate_chapter", self.generate_chapter)
        graph_builder.add_node("human_feedback_on_chapter", self.human_feedback_on_chapter)
        graph_builder.add_node("read_chapter_feedback", self.read_chapter_feedback)
        graph_builder.add_node("regenerate_chapter", self.regenerate_chapter)
        graph_builder.add_node("generate_chapter_summary_and_update_db", self.generate_chapter_summary_and_update_db)
        graph_builder.add_node("ask_for_next_chapter", self.ask_for_next_chapter)
        graph_builder.add_node("ask_for_final_review", self.ask_for_final_review)
        graph_builder.add_node("read_final_review_notes", self.read_final_review_notes)
        graph_builder.add_node("perform_final_revision", self.perform_final_revision)
        graph_builder.add_node("finish", self.finish_node)

        # Build the graph
        graph_builder.add_edge(START, "read_local_excel")
        graph_builder.add_conditional_edges(
            "read_local_excel", self.decide_to_generate_outline,
            {"continue": "generate_outline", "stop": "finish"}
        )
        graph_builder.add_edge("generate_outline", "write_outline_to_excel")
        graph_builder.add_edge("write_outline_to_excel", "human_feedback_interrupt")
        graph_builder.add_edge("human_feedback_interrupt", "read_feedback_from_excel")
        graph_builder.add_edge("read_feedback_from_excel", "generate_chapter")
        
        graph_builder.add_edge("generate_chapter", "human_feedback_on_chapter")
        graph_builder.add_edge("human_feedback_on_chapter", "read_chapter_feedback")
        graph_builder.add_conditional_edges(
            "read_chapter_feedback", self.decide_to_regenerate_chapter,
            {"continue": "regenerate_chapter", "skip": "generate_chapter_summary_and_update_db"}
        )

        graph_builder.add_edge("regenerate_chapter", "generate_chapter_summary_and_update_db")
        graph_builder.add_edge("generate_chapter_summary_and_update_db", "ask_for_next_chapter")
        
        graph_builder.add_conditional_edges(
            "ask_for_next_chapter", self.should_generate_next_chapter,
            {"continue": "generate_chapter", "stop": "ask_for_final_review"}
        )
        graph_builder.add_edge("ask_for_final_review", "read_final_review_notes")
        graph_builder.add_conditional_edges(
            "read_final_review_notes", self.decide_on_final_revision,
            {"revise": "perform_final_revision", "finish": "finish"}
        )
        graph_builder.add_edge("perform_final_revision", "finish")
        
        graph_builder.add_edge("finish", END)

        self.checkpointer = InMemorySaver()
        self.graph = graph_builder.compile(checkpointer=self.checkpointer)

    def _sanitize_filename(self, filename: str) -> str:
        return re.sub(r'[\\/*?:"<>|]', "", filename)

    def _sync_read_excel(self):
        try:
            workbook = openpyxl.load_workbook(self.excel_filename)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            data_row = [cell.value for cell in sheet[2]] if sheet.max_row > 1 else []
            if not data_row or not any(data_row): return {}
            data = dict(zip(headers, data_row))
            return data
        except FileNotFoundError:
            raise FileNotFoundError(f"Error: The file '{self.excel_filename}' was not found.")
        except Exception as e:
            raise RuntimeError(f"An error occurred while reading the Excel file: {e}")

    def _sync_write_excel(self, column_name: str, content: str, row_index: int = 2):
        try:
            workbook = openpyxl.load_workbook(self.excel_filename)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            col_idx = -1
            if column_name in headers:
                col_idx = headers.index(column_name) + 1
            else:
                col_idx = len(headers) + 1
                sheet.cell(row=1, column=col_idx, value=column_name)
            sheet.cell(row=row_index, column=col_idx, value=content)
            workbook.save(self.excel_filename)
        except Exception as e:
            raise RuntimeError(f"An error occurred while writing to the Excel file: {e}")
    
    def _sync_write_to_text_file(self, filename: str, content: str):
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(content)
        except Exception as e:
            raise RuntimeError(f"An error occurred while writing to the text file: {e}")

    def _sync_append_to_text_file(self, filename: str, content: str):
        try:
            with open(filename, 'a', encoding='utf-8') as f:
                f.write(content)
        except Exception as e:
            raise RuntimeError(f"An error occurred while appending to the text file: {e}")

    def _sync_get_user_input(self, prompt: str) -> str:
        return input(prompt)

    def _send_email(self, subject, body, attachment_content=None, attachment_filename=None):
        if not all([self.sender_email, self.email_password, self.receiver_email]):
            print("Email credentials not found in .env file. Skipping email notification.")
            return

        msg = MIMEMultipart()
        msg['From'] = self.sender_email
        msg['To'] = self.receiver_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        if attachment_content and attachment_filename:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment_content.encode('utf-8'))
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{attachment_filename}"')
            msg.attach(part)

        try:
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
                server.login(self.sender_email, self.email_password)
                server.send_message(msg)
                print("Email notification sent successfully.")
        except Exception as e:
            print(f"Failed to send email: {e}")

    async def read_local_excel(self, state: State):
        print("--- Reading from local Excel file ---")
        data = await asyncio.to_thread(self._sync_read_excel)
        title = data.get('title')
        notes = data.get('notes_before_outline')
        sanitized_title = self._sanitize_filename(title) if title else ""
        print(f"Title: {title or 'Not Found'}\nNotes: {notes or 'Not Found'}\n")
        return {
            "title": title, "notes_before_outline": notes, "chapters": [], 
            "chapter_summaries": [], 
            "current_chapter_number": 1, "sanitized_title": sanitized_title
        }

    def decide_to_generate_outline(self, state: State):
        return "continue" if state.get("title") and state.get("notes_before_outline") else "stop"

    async def generate_outline_node(self, state: State):
        print("--- Generating Outline ---")
        prompt = ChatPromptTemplate.from_messages([
            ("system", "You are an expert book author. Generate a detailed outline for a book."),
            ("human", "Title: {title}\nInitial Notes: {notes_before_outline}")
        ])
        chain = prompt | self.llm
        outline = (await chain.ainvoke(state)).content
        print(f"--- Generated Outline ---")
        return {"outline": outline}

    async def write_outline_to_excel(self, state: State):
        print("--- Writing outline to Excel ---")
        await asyncio.to_thread(self._sync_write_excel, "outline_generated_by_llm", state["outline"])
        print("Outline successfully written to 'outline_generated_by_llm' column.\n")
        subject = f"Outline Generated for '{state['title']}'"
        body = "The book outline has been generated. Please see the attached file."
        await asyncio.to_thread(self._send_email, subject, body, state["outline"], f"{state['sanitized_title']}_outline.txt")
        return {}

    async def human_feedback_interrupt(self, state: State):
        print("--- Waiting for Human Feedback on Outline ---")
        prompt = (
            "Outline written to Excel. Please review it.\n"
            "If you have improvements, add them to the 'notes_on_outline_after' column.\n"
            "Press Enter here to continue once you are done.\n"
        )
        await asyncio.to_thread(self._sync_get_user_input, prompt)
        return {}

    async def read_feedback_from_excel(self, state: State):
        print("--- Reading outline feedback from Excel ---")
        data = await asyncio.to_thread(self._sync_read_excel)
        feedback = data.get('notes_on_outline_after')
        print(f"Feedback received: {feedback}\n" if feedback else "No new feedback provided.\n")
        return {"notes_on_outline_after": feedback}

    async def generate_chapter(self, state: State):
        chapter_num = state['current_chapter_number']
        print(f"--- Generating Chapter {chapter_num} ---")
        prompt_template = ChatPromptTemplate.from_messages([
            ("system", "You are a master storyteller. Write a single chapter for a book based on the provided details. Start with a chapter title."),
            ("human", "Book Title: {title}\nFull Book Outline:\n{outline}\nAuthor's Feedback on Outline:\n{notes_on_outline_after}\n\nPlease write the full content for Chapter {chapter_num} now.")
        ])
        chain = prompt_template | self.llm
        chapter_content = (await chain.ainvoke({
            "title": state["title"], "outline": state["outline"],
            "notes_on_outline_after": state.get("notes_on_outline_after") or "No feedback provided.",
            "chapter_num": chapter_num
        })).content
        print(f"--- Chapter {chapter_num} Generated ---")
        
        book_filename = f"{state['sanitized_title']}.txt"
        chapter_header = f"\n\n---\n\nChapter {chapter_num}\n\n---\n\n"
        await asyncio.to_thread(self._sync_append_to_text_file, book_filename, chapter_header + chapter_content)
        print(f"Chapter {chapter_num} content appended to '{book_filename}'\n")

        subject = f"Chapter {chapter_num} Generated for '{state['title']}'"
        body = f"Chapter {chapter_num} has been generated. Please see the attached file."
        await asyncio.to_thread(self._send_email, subject, body, chapter_content, f"{state['sanitized_title']}_chapter_{chapter_num}.txt")

        return {"chapters": state.get("chapters", []) + [chapter_content]}

    async def human_feedback_on_chapter(self, state: State):
        chapter_num = state['current_chapter_number']
        print(f"--- Waiting for Feedback on Chapter {chapter_num} ---")
        prompt = (
            f"Chapter {chapter_num} has been appended to {state['sanitized_title']}.txt.\n"
            "To suggest changes, add notes to the 'chapter_notes' column in the Excel file.\n"
            "Press Enter here to continue once you are done.\n"
        )
        await asyncio.to_thread(self._sync_get_user_input, prompt)
        return {}

    async def read_chapter_feedback(self, state: State):
        print("--- Reading chapter feedback from Excel ---")
        data = await asyncio.to_thread(self._sync_read_excel)
        feedback = data.get('chapter_notes')
        print(f"Chapter feedback received: {feedback}\n" if feedback else "No chapter feedback provided.\n")
        return {"chapter_feedback": feedback}

    def decide_to_regenerate_chapter(self, state: State):
        return "continue" if state.get("chapter_feedback") else "skip"

    async def regenerate_chapter(self, state: State):
        chapter_num = state['current_chapter_number']
        print(f"--- Regenerating Chapter {chapter_num} based on feedback ---")
        
        prompt = ChatPromptTemplate.from_messages([
            ("system", "You are a master storyteller. Revise a book chapter based on the author's feedback."),
            ("human", "Original Chapter Content:\n{original_chapter}\n\nAuthor's Feedback for Revision:\n{feedback}\n\nPlease rewrite the entire chapter, incorporating the feedback.")
        ])
        chain = prompt | self.llm
        original_chapter = state['chapters'][-1]
        
        revised_content = (await chain.ainvoke({"original_chapter": original_chapter, "feedback": state['chapter_feedback']})).content
        updated_chapters = state['chapters'][:-1] + [revised_content]

        all_chapters_text = "".join(
            f"\n\n---\n\nChapter {i+1}\n\n---\n\n{content}" for i, content in enumerate(updated_chapters)
        ).strip()
        
        book_filename = f"{state['sanitized_title']}.txt"
        await asyncio.to_thread(self._sync_write_to_text_file, book_filename, all_chapters_text)
        print(f"Book file '{book_filename}' updated with revised Chapter {chapter_num}.\n")

        await asyncio.to_thread(self._sync_write_excel, "chapter_notes", "")
        print("Processed chapter feedback and cleared 'chapter_notes' in Excel.\n")

        return {"chapters": updated_chapters, "chapter_feedback": None}

    async def generate_chapter_summary_and_update_db(self, state: State):
        current_chapter_num = state['current_chapter_number']
        print(f"--- Generating Summary for Chapter {current_chapter_num} ---")
        
        latest_chapter_content = state["chapters"][-1]

        prompt = ChatPromptTemplate.from_messages([
            ("system", "You are an expert summarizer. Create a concise summary of the provided book chapter."),
            ("human", "Chapter content:\n\n{chapter_content}\n\nPlease provide a summary for this chapter.")
        ])
        chain = prompt | self.llm
        summary_content = (await chain.ainvoke({"chapter_content": latest_chapter_content})).content

        new_summary = {
            "chapter": current_chapter_num,
            "summary": summary_content
        }
        print(f"Summary for Chapter {current_chapter_num} generated.\n")

        chapter_summaries = state.get("chapter_summaries", [])
        summary_exists = False
        for i, summary in enumerate(chapter_summaries):
            if summary.get("chapter") == current_chapter_num:
                chapter_summaries[i] = new_summary
                summary_exists = True
                break
        
        if not summary_exists:
            chapter_summaries.append(new_summary)

        print(f"--- Updating MongoDB for book: '{state['title']}' ---")
        try:
            await self.db.update_book_chapters(
                book_title=state['title'],
                chapters_data=chapter_summaries
            )
            print("MongoDB update successful.\n")
        except Exception as e:
            print(f"An error occurred while updating MongoDB: {e}\n")

        return {"chapter_summaries": chapter_summaries}

    async def ask_for_next_chapter(self, state: State):
        next_chapter_num = state['current_chapter_number'] + 1
        prompt = f"Do you want to generate Chapter {next_chapter_num}? (yes/no): "
        decision = await asyncio.to_thread(self._sync_get_user_input, prompt)
        return {"user_decision": decision.lower(), "current_chapter_number": next_chapter_num}

    def should_generate_next_chapter(self, state: State):
        return "continue" if state.get("user_decision") == "yes" else "stop"

    async def ask_for_final_review(self, state: State):
        print("--- Final Review Stage ---")
        prompt = (
            f"The book '{state['title']}' has been fully generated.\n"
            "To perform a final revision, please add your notes to the 'final_review_notes' column in the Excel file.\n"
            "IMPORTANT: Then, set the 'final_review_notes_status' column to 'pending' to trigger the revision.\n"
            "If no final changes are needed, you can just press Enter to finish.\n"
        )
        await asyncio.to_thread(self._sync_get_user_input, prompt)
        return {}
    
    async def read_final_review_notes(self, state: State):
        print("--- Reading final review notes from Excel ---")
        data = await asyncio.to_thread(self._sync_read_excel)
        status = data.get('final_review_notes_status')
        notes = data.get('final_review_notes')
        
        if status == 'pending' and notes:
            print(f"Final review notes received: {notes}\n")
        else:
            print("No pending final review notes found. Proceeding to finish.\n")
        return {"final_review_notes_status": status, "final_review_notes": notes}

    def decide_on_final_revision(self, state: State):
        status = state.get("final_review_notes_status", "").lower()
        notes = state.get("final_review_notes")
        return "revise" if status == "pending" and notes else "finish"

    async def perform_final_revision(self, state: State):
        print(f"--- Performing Final Revision for '{state['title']}' ---")
        full_book_content = "\n\n".join(state['chapters'])

        prompt = ChatPromptTemplate.from_messages([
            ("system", "You are a master editor. Your task is to perform a final, comprehensive revision of an entire book based on the author's concluding notes. Polish the manuscript, improve flow, and correct any inconsistencies."),
            ("human", "Full book content:\n\n{full_book_content}\n\nAuthor's final revision notes:\n{final_notes}\n\nPlease provide the complete, final version of the entire book.")
        ])
        chain = prompt | self.llm
        final_book_content = (await chain.ainvoke({
            "full_book_content": full_book_content, 
            "final_notes": state['final_review_notes']
        })).content
        
        final_filename = f"{state['sanitized_title']}_FINAL.txt"
        await asyncio.to_thread(self._sync_write_to_text_file, final_filename, final_book_content)
        print(f"Final revised version of the book has been saved to '{final_filename}'\n")

        await asyncio.to_thread(self._sync_write_excel, "final_review_notes_status", "completed")
        await asyncio.to_thread(self._sync_write_excel, "final_review_notes", "")
        
        subject = f"Final Version of '{state['title']}' is Ready"
        body = "The final, revised version of your book is complete. Please see the attached file."
        await asyncio.to_thread(self._send_email, subject, body, final_book_content, final_filename)

        return {}

    async def finish_node(self, state: State):
        print("--- Workflow Finished ---")
        return {}

async def main():
    book_gen = BookGeneration("Book Generation.xlsx")
    config = {"configurable": {"thread_id": "interactive-book-gen-thread-v5"}} 
    initial_state = {"messages": [HumanMessage(content="Start")]}
    
    try:
        async for event in book_gen.graph.astream(initial_state, config=config):
            for key, value in event.items():
                print(f"Node: {key}")
                print("---")
    except (FileNotFoundError, ValueError, RuntimeError) as e:
        print(f"\nExecution failed: {e}")

if __name__ == "__main__":
    asyncio.run(main())